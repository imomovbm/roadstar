from django.shortcuts import render, redirect, get_object_or_404
import requests
import time
import hashlib
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from openpyxl import load_workbook
from django.views.decorators.csrf import csrf_exempt
import json
from django.conf import settings
from django.http import JsonResponse, HttpResponse
from pathlib import Path
import re

def parse_number(value):
    try:
        return float(value.replace(',', ''))
    except:
        return 0

def doi(request):
    return(HttpResponse("Shaxi style bolmasin eee"))

@login_required
def index(request):
    return render(request, "sms_gate/sms.html")

@login_required
def send_sms(request):
    if request.method == 'POST':
        try:
            wb = load_workbook(filename=request.FILES['excel_file'], read_only=True, data_only=True)
            ws = wb.active
            max_row = ws.max_row
            clients = []
            
            pattern = r"^№.*\d.* от \d{2}\.\d{2}\.\d{4}$"


            for row_num in range(1, max_row + 1):
                contract_id = str(ws[f'A{row_num}'].value).strip()
                
                if not re.match(pattern, contract_id):
                    continue  # Skip rows not matching the pattern

                debt = ws[f'J{row_num}'].value
                advance = ws[f'K{row_num}'].value

                if debt is None and advance is None:
                    continue

                debt_val = parse_number(str(debt)) if debt else 0
                advance_val = parse_number(str(advance)) if advance else 0

                if debt_val < 1000000 and advance_val < 1000000:
                    continue  # Skip if both values are less than 100,000
                contract_id = contract_id.replace('от', 'raqamli')
                client = (contract_id, debt_val, advance_val)
                clients.append(client)

                # You can integrate SMS sending here
                # send_sms_ibnux(phone, f"Shartnoma {contract_id}: Qarz: {debt_val}, Avans: {advance_val} so'm")

            messages.success(request, f"{len(clients)} ta SMS jo'natildi!")

        except Exception as e:
            messages.error(request, f"Xatolik: {str(e)}")

    return render(request, "sms_gate/sms.html", {
        "clients": clients})

# Create your views here.
def send_sms_ibnux(phone, message):
    device_id = "f_aTAkv6EnI:APA91bH0IeR1uV8xub0p1uclhl0IoT-5tJ5MXqM6KD2jiI3VO74r1zGOeaECDe9Y_jlu1Z86n2vaUAZjXNDfYWb56NZTzwi__YUGlPdCqhRp8OHS1dUtpvA"
    secret_key = "d5ece1d4-9eee-43dc-ad0e-e488cbe02285"
    now = str(int(time.time()))
    hashed_secret = hashlib.md5((secret_key + now).encode()).hexdigest()
    
    payload = {
        "deviceID": device_id,
        "secret": hashed_secret,
        "time": now,
        "to": phone,
        "text": message
    }

    response = requests.post("https://sms.ibnux.net/", data=payload)
    return response.status_code == 200